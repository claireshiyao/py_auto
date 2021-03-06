"""
1、打开表单

2、读取标题 头部

3、读取所有的数据 (类型是？？)

4、指定单元格写入数据（使用静态方法，不要使用实例方法）
"""

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler():
    """操作 Excel """

    def __init__(self, file):
        """初始化函数"""
        self.file = file
        # self.wb = load_workbook(file)

    def open_sheet(self, name) -> Worksheet:
        """打开表单.
        在函数或者方法的后面 加 -> 类型：表示此函数返回值是一个 这样的类型
        函数注解。
        """
        wb = load_workbook(self.file)
        sheet = wb[name]
        wb.close()
        return sheet

    def header(self, sheet_name):
        """获取表单的表头"""
        sheet = self.open_sheet(sheet_name)
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    def read(self, sheet_name):
        """读取所有的数据"""
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows) # 不加List是生成器对象generator object

        # 获取标题
        data = []
        for row in rows[1:]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
                # 列表转成字典：要和 header去 zip
                data_dict = dict(zip(self.header(sheet_name), row_data))
            data.append(data_dict)
        return data

    @staticmethod
    def write(file, sheet_name, row, column, data):
        """写入 Excel 数据"""
        wb = load_workbook(file)
        sheet = wb[sheet_name]
        # 修改单元格
        sheet.cell(row, column).value = data
        # 保存
        wb.save(file)
        # 关闭
        wb.close()


if __name__ == '__main__':
    excel = ExcelHandler(r'd:\cases.xlsx')
    excel.write(r'd:\cases.xlsx', 'Sheet1', 5, 1, 'data_value')

