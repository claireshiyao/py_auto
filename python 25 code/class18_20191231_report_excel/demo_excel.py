#!/usr/bin/env python3
#-*- coding:utf-8 -*-
# email: wagyu2016@163.com
# wechat: shoubian01
# author: 王雨泽

"""
保存测试数据

Excel 的基础用法。手工如何操作 Excel， python 学习手工如何操作Excel

手工操作 Excel 的流程
1,  打开Excel 文件 （路径 + 文件名）
2， 获取表单
3， 使用行号，和列好去确定需要读取的数据
4， 关闭文件

python 操作 Excel, 工具
- openpyxl,  支持 xlsx 新型格式的读写， 读取速度还可以。
- tablib， 支持多种格式读写。 xlsx, xls, csv, json, yaml, html, pd
- xlrd, 经典的 Excel 读取库，
- xlwt,
- pandas, 功能强，太臃肿了

安装： pip install openpyxl

获取表单：
1，wb.active,  被选中，被激活
2，通过索引， wb.worksheets[索引]
3，通过sheet名字， wb['sheet_name']

"""
import openpyxl

# 读取 Excel 文件夹
# 读取文件之前一定要关闭该文件
# windows下面的路径有反斜杠
# 得到一个 Workbook 对象
wb = openpyxl.load_workbook(r'd:\cases.xlsx')
print(wb)
# 不直接去获取_sheets 属性，称为私有属性。
print(wb._sheets) # self._sheets
print(wb.sheetnames) #['Sheet1', 'Sheet2', 'Sheet3']

# active 是表示被激活，被选择的 sheet
active_sheet = wb.active

# sheetnames 和 _sheets 有什么区别？？
# sheetnames 列表当中存储的是字符串， _sheets 里面存储是对象。

#
# 获取所有表单的正确用法
# wb.worksheets
# # 获取某一个表单，1， 通过索引去获取
sheet = wb.worksheets[0]

# 第二种方式：通过表单名称获取， Sheet1
# sheet = wb.get_sheet_by_name("Sheet1")   # 过时
# print(sheet)
# 正规用法

# sheet = wb['Sheet1']

# print(sheet)
# Pycharm 不支持 sheet.属性的提示。

# 读取单个单元格, 行和列
# 行和列是从 1 开始的，不是 python 当中从 0 开始
# cell = sheet.cell(1, 2)
# print(cell)
#
# # 获取 cell 的值, 很多同学会忘记加上 value
# print(cell.value)
#
# # 获取某一行
# print(sheet[1])
# # 获取值：
# for column in sheet[1]:
#     print(column.value)
#
#
# # 获取某一列
# print(sheet['A'])
#
# # 如何获取多行。 1 到 3 行, 第三行是包含的。
# print(sheet[1:3])

# 获取所有行的数据
print(sheet.rows)  # <generator object Worksheet._cells_by_row at 0x000001729FD54138>
total_data = list(sheet.rows)
print(total_data)
# [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>), (<Cell 'Sheet1'.A4>, <Cell 'Sheet1'.B4>)]

for row in total_data:
    for cell in row:
        print(cell.value)

# 写称类

# 写入。一个单元格，
# 保存， save("文件名称")
wb.save(r'd:\cases.xlsx')

# 关闭
wb.close()
