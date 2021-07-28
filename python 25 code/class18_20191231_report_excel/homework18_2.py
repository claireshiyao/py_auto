# author by claire

# excel类封装需要提供以下功能：
# 1、打开表单
# 2、读取标题
# 3、读取所有的数据
# 4、指定单元格写入数据（使用静态方法，不要使用实例方法）

import openpyxl
class Excel:
    # def __init__(self, excel_path, sheetname):
    #     self.excel_path = excel_path
    #     self.sheetname = sheetname

    # @staticmethod
    # def open_excel(excel_path):
    #     """打开excel文件"""
    #     openpyxl.load_workbook(excel_path)

    def get_sheet(self, excel_path, sheetname):
        """打开表单"""
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.get_sheet_by_name(sheetname)
        return sheet

    def read_title(self, excel_path, sheetname):
        """读取标题"""
        sheet = self.get_sheet(excel_path, sheetname)
        for title in sheet[1]:
            return title.value

    def read_datas(self, excel_path, sheetname):
        """获取所有的数据"""
        sheet = self.get_sheet(excel_path, sheetname)
        total_data = list(sheet.rows)
        for row in total_data:
            for cell in row:
                return cell.value

    @ staticmethod
    def write_data(self, excel_path, sheetname, row, column, new_value):
        """指定单元格写入数据"""
        sheet = self.get_sheet(excel_path, sheetname)
        cell = sheet.cell(row, column)
        cell.value = new_value

    def save_excel(self, excel_path):
        """保存数据，存/另存为文件"""
        self.open_excel(excel_path).save()

    def close_excel(self, excel_path):
        """关闭文件"""
        self.open_excel(excel_path).close()

if __name__ == '__main__':
    excel = Excel()
    excel_path = r'd:\test_cases.xlsx'
    sheet_name = "Sheet1"
    #读取标题
    print(excel.read_title(excel_path, sheet_name))
    #获取所有数据
    print(excel.read_datas(excel_path, sheet_name))
    #写入数据
    print(excel.write_data(excel_path, sheet_name, 3, 1, 'xsy3'))
    #保存
    excel.save_excel(excel_path)
    #关闭
    excel.close_excel(excel_path)


