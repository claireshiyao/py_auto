# author by claire

# excel类封装需要提供以下功能：
# 1、打开表单
# 2、读取标题
# 3、读取所有的数据
# 4、指定单元格写入数据（使用静态方法，不要使用实例方法）

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class Excel:
    def __init__(self, excel_path):
        self.excel_path = excel_path

    def open_sheet(self, sheetname) -> Worksheet:
        """打开表单"""
        wb = openpyxl.load_workbook(self.excel_path)
        sheet = wb[sheetname]
        wb.close()
        return sheet

    def read_title(self, sheet_name):
        """读取标题"""
        sheet = self.open_sheet(sheet_name)
        title = []
        for i in sheet[1]:
            title.append(i.value)
        return title

    def read_datas(self, sheet_name):
        """获取所有的数据"""
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows)

        data = []
        for row in rows[1:]:
            row_data = []
            for cell in row:
                # print(cell.value)
                row_data.append(cell.value)
                # 列表转成字典
                data_dict = dict(zip(self.read_title(sheet_name), row_data))
            data.append(data_dict)
        return data

    @ staticmethod
    def write_data(file, sheet_name, row, column, new_value):
        """指定单元格写入数据"""
        # sheet = self.open_sheet(sheet_name)  # 静态方法不能这样调用
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        cell = sheet.cell(row, column)
        cell.value = new_value
        wb.save(file)
        wb.close()

if __name__ == '__main__':
    excel = Excel(r'd:\test_cases.xlsx')
    # sheet_new = excel.open_sheet('sheet1')

    #读取标题
    print(excel.read_title('Sheet1'))
    #获取所有数据
    print(excel.read_datas('Sheet1'))
    #写入数据
    print(excel.write_data(r'd:\test_cases.xlsx', 'Sheet1', 4, 1, 'xsy3'))


